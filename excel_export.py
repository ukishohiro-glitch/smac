# -*- coding: utf-8 -*-
# excel_export.py - テンプレ罫線/結合を維持した転記（見積書0/1〜5対応）
# 要件:
#   見積書0: J1/J3/A6/A7/A8/B17 ヘッダ、21〜44 行に間口合計を連続記載（A=カーテン品名 / F=1 / G=式 / H=合計）
#            直後に「運賃・梱包」を空行なしで追記（price>0 の場合）
#   見積書1〜5: 11行目は触らない / 12〜44 行、A=品名 / F=数量 / G=単位 / H=単価
#            入力順、1間口33行以内はページ跨ぎ禁止、33行超はページ跨ぎ可、5ページ超はエラー

from typing import List, Dict, Any
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# ---------- ユーティリティ ----------
def _to_int(x) -> int:
    try:
        return int(x)
    except Exception:
        try:
            return int(float(x))
        except Exception:
            return 0

def _first_curtain_name(items: List[Dict[str, Any]]) -> str:
    if not items: return ""
    keywords = ["S・MAC", "エア・セーブ", "ME", "カーテン"]
    for it in items:
        nm = str(it.get("品名") or it.get("name") or it.get("product") or "")
        for kw in keywords:
            if kw in nm:
                return nm
    return str(items[0].get("品名") or items[0].get("name") or items[0].get("product") or "")

def _opening_total(items: List[Dict[str, Any]]) -> int:
    tot = 0
    for it in items:
        q = _to_int(it.get("数量") or it.get("qty") or 0)
        p = _to_int(it.get("単価") or it.get("unit_price") or 0)
        tot += q * p
    return tot

def _normalize_openings(overall_items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    overall_items が間口ごとの構造ならそのまま、平坦なら1間口として解釈。
    """
    if not overall_items:
        return []
    if isinstance(overall_items[0], dict) and "items" in overall_items[0]:
        return [
            {"name": op.get("name") or "", "items": list(op.get("items") or []), "teikeiku": list(op.get("teikeiku") or [])}
            for op in overall_items
        ]
    return [{"name": "", "items": overall_items, "teikeiku": []}]

# ---------- 見積書0（ヘッダ/合計） ----------
def _write_header(ws0: Worksheet, header: Dict[str, Any]) -> None:
    ws0["J1"].value = header.get("見積番号", "")
    ws0["J3"].value = header.get("作成日", "")
    ws0["A6"].value = header.get("得意先名", "")
    b = header.get("支店名", ""); o = header.get("営業所名", "")
    ws0["A7"].value = (str(b) + " " + str(o)).strip()
    ws0["A8"].value = header.get("担当者名", "")
    ws0["B17"].value = header.get("物件名", "")

def _write_sheet0_totals(ws0: Worksheet, openings: List[Dict[str, Any]], header: Dict[str, Any]) -> None:
    """
    21〜44行に間口合計を連続記載、直後に運賃・梱包（金額>0なら）。
    """
    row = 21
    for op in openings:
        items = op.get("items", [])
        ws0.cell(row=row, column=1).value = _first_curtain_name(items)  # A
        ws0.cell(row=row, column=6).value = 1                            # F
        ws0.cell(row=row, column=7).value = "式"                         # G
        ws0.cell(row=row, column=8).value = _opening_total(items)        # H
        row += 1
        if row > 44:
            raise ValueError("見積書0の行数が 21〜44 を超えました。")
    ship = header.get("shipping")
    if ship and _to_int(ship.get("price", 0)) > 0:
        label = str(ship.get("label", "運賃・梱包")) + str(ship.get("option", ""))
        ws0.cell(row=row, column=1).value = label
        ws0.cell(row=row, column=6).value = 1
        ws0.cell(row=row, column=7).value = "式"
        ws0.cell(row=row, column=8).value = _to_int(ship.get("price", 0))

# ---------- 明細（見積書1〜5） ----------
def _write_detail_row(ws: Worksheet, row: int, it: Dict[str, Any]) -> None:
    ws.cell(row=row, column=1).value = str(it.get("品名") or it.get("name") or it.get("product") or "")
    ws.cell(row=row, column=6).value = it.get("数量") if it.get("数量") is not None else it.get("qty")
    ws.cell(row=row, column=7).value = it.get("単位") if it.get("単位") is not None else it.get("unit")
    ws.cell(row=row, column=8).value = it.get("単価") if it.get("単価") is not None else it.get("unit_price")

def _write_details(wb, openings: List[Dict[str, Any]], detail_sheets: List[str], start_row: int = 12, end_row: int = 44) -> None:
    rows_per_page = end_row - start_row + 1  # 33
    page_max = len(detail_sheets)
    page_idx = 0
    ws = wb[detail_sheets[page_idx]]
    r = start_row

    def ensure_next_page():
        nonlocal page_idx, ws, r
        page_idx += 1
        if page_idx >= page_max:
            raise ValueError("5ページを超えました。生成不可")
        ws = wb[detail_sheets[page_idx]]
        r = start_row

    for op in openings:
        items = list(op.get("items", []))
        teikeiku = list(op.get("teikeiku", []))
        need = len(items) + len(teikeiku)

        if need <= rows_per_page:
            # 1間口33行以内 → ページ跨ぎ禁止（丸ごと）
            if r + need - 1 > end_row:
                ensure_next_page()
            for it in items:
                _write_detail_row(ws, r, it); r += 1
            for s in teikeiku:
                ws.cell(row=r, column=1).value = str(s); r += 1
        else:
            # 33行超 → 跨ぎ可
            for it in items:
                if r > end_row: ensure_next_page()
                _write_detail_row(ws, r, it); r += 1
            for s in teikeiku:
                if r > end_row: ensure_next_page()
                ws.cell(row=r, column=1).value = str(s); r += 1

# ---------- エクスポート ----------
def export_quotation_book_preserve(out_path, header, overall_items, *, template_path, header_sheet="見積書0",
                                   detail_sheets=None, start_row=12, end_row=44):
    if detail_sheets is None:
        detail_sheets = [f"見積書{i}" for i in range(1, 6)]
    wb = load_workbook(template_path)
    if header_sheet not in wb.sheetnames:
        raise ValueError(f"テンプレートに {header_sheet} がありません。")
    for s in detail_sheets:
        if s not in wb.sheetnames:
            raise ValueError(f"テンプレートに {s} がありません。")

    ws0 = wb[header_sheet]
    openings = _normalize_openings(overall_items)

    _write_header(ws0, header)
    _write_sheet0_totals(ws0, openings, header)
    _write_details(wb, openings, detail_sheets, start_row=start_row, end_row=end_row)

    wb.save(out_path)

def export_detail_xlsx_preserve(out_path, header, overall_items, *, template_path, ws_name="お見積書（明細）", start_row=12, max_rows=33):
    """旧テンプレ互換（単一シート）。可能な範囲で出力して保険にする。"""
    wb = load_workbook(template_path)
    if ws_name not in wb.sheetnames:
        raise ValueError(f"テンプレートに {ws_name} がありません。")
    ws = wb[ws_name]

    # 最低限のヘッダ（位置はテンプレ依存のため共通部のみ）
    ws["J1"].value = header.get("見積番号", "")
    ws["J3"].value = header.get("作成日", "")

    page = 0
    row = start_row
    rows_per_page = max_rows
    pages_limit = 5

    def ensure_next_page():
        nonlocal page, row
        page += 1
        if page >= pages_limit:
            raise ValueError("5ページを超えました。生成不可")
        row = start_row + page * rows_per_page

    openings = _normalize_openings(overall_items)
    for op in openings:
        items = list(op.get("items", []))
        teikeiku = list(op.get("teikeiku", []))
        need = len(items) + len(teikeiku)

        # 33行以内の間口はページ跨ぎ禁止（丸ごと次ページへ送る）
        cur_used = (row - (start_row + page * rows_per_page))
        if need <= rows_per_page and cur_used + need > rows_per_page:
            ensure_next_page()

        for it in items:
            if (row - (start_row + page * rows_per_page)) >= rows_per_page:
                ensure_next_page()
            _write_detail_row(ws, row, it)
            row += 1
        for s in teikeiku:
            if (row - (start_row + page * rows_per_page)) >= rows_per_page:
                ensure_next_page()
            ws.cell(row=row, column=1).value = str(s)
            row += 1

    wb.save(out_path)
