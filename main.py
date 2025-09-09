# -*- coding: utf-8 -*-
# main.py — 本番対応 完全版（更新：寸法行/定型句/空行/運賃表記）
# 変更点（今回）:
#  1) 見積サマリ＆Excel明細の表示仕様
#     - S・MAC: 品名行の下に「間口寸法」行＆「カーテン寸法」行を追加。
#               数量/単位/単価/小計は「カーテン寸法」行に記載（品名行は空欄、間口寸法行も空欄）。
#     - エア・セーブ: 品名行の下に「間口寸法」行を追加。
#               数量/単位/単価/小計は「間口寸法」行に記載（品名行は空欄）。
#  2) 定型句は「間口ごと」に選択（既存仕様のままですが内部キーを厳密化）。
#  3) 複数間口の場合、Excel明細で間口ブロック間に1行の空行を自動挿入
#     （定型句があればその直下に空行）。
#  4) 運賃ラベルを「路線便（時間指定不可）」「現場搬入（時間指定可）」に変更
#     → サマリ(見積書0)の表記も同一。

import os, os.path as osp, secrets, math, re, unicodedata
from datetime import datetime, date
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_to_tuple
from pathlib import Path

# ===== パス =====
APP_DIR = Path(__file__).parent
TEMPLATE_BOOK = APP_DIR / "お見積書（明細）.xlsx"    # テンプレファイル名
MASTER_BOOK   = APP_DIR / "master.xlsx"

# ===== ユーティリティ =====
def normalize_string(s):
    if isinstance(s, str):
        s = s.replace("\u3000", " ")
        s = unicodedata.normalize("NFKC", s)
        s = re.sub(r"\s+", " ", s.strip())
    return s

def norm_cols(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty: return pd.DataFrame()
    df = df.copy()
    df.columns = [normalize_string(c) for c in df.columns]
    for c in df.columns:
        if df[c].dtype == object:
            df[c] = df[c].map(normalize_string)
    return df

NUM_RE = re.compile(r"[-+]?\d[\d,]*\.?\d*")
def parse_float(x):
    if x is None: return None
    m = NUM_RE.search(str(x))
    if not m: return None
    try: return float(m.group(0).replace(",", ""))
    except: return None

def parse_length_mm(x):
    if x is None: return None
    s = normalize_string(str(x)).lower()
    m_m = re.search(r"(\d[\d,]*\.?\d*)\s*m(?!m)", s)
    if m_m:
        return int(math.ceil(float(m_m.group(1).replace(",", ""))*1000))
    v = parse_float(s)
    if v is None: return None
    if "mm" in s: return int(round(v))
    return int(round(v if v>=1000 else v*1000))

def ceil100(x: float) -> int: return int(math.ceil(x/100.0)*100)

def pick_col(df: pd.DataFrame, candidates) -> str | None:
    for c in candidates:
        if c in df.columns: return c
    norm = {normalize_string(c): c for c in df.columns}
    for c in candidates:
        if c in norm: return norm[c]
    return None

def read_xlsx(path: str, sheet: str | None = None) -> pd.DataFrame:
    try:
        df = pd.read_excel(path, sheet_name=sheet) if sheet else pd.read_excel(path)
        return norm_cols(df)
    except Exception:
        return pd.DataFrame()

# ===== マスタ読込 =====
def load_master():
    def R(sn, add_cols=None):
        try:
            df = pd.read_excel(MASTER_BOOK, sheet_name=sn); df = norm_cols(df)
            if add_cols:
                for c in add_cols:
                    if c not in df.columns: df[c] = ""
            return df
        except Exception:
            return pd.DataFrame()

    df_clients  = R("得意先一覧")
    df_curtain  = R("カーテン", ["大分類","中分類","小分類"])
    df_perf     = R("カーテン性能", ["中分類","性能"])
    df_ma       = R("MA型単価表")
    df_mb_tbl   = R("MB型単価表")
    df_mc       = R("MC型単価表")
    df_me_curt  = read_xlsx(MASTER_BOOK, "ME型単価表")
    df_me_motor = read_xlsx(MASTER_BOOK, "ME型OP")  # MEの駆動・センサー等
    df_op       = R("OP", ["OP名称","金額","方向"])
    df_gap      = R("隙間シート")
    parts_sheets = ["カーテンレール","取手付間仕切ポール","中間ポール","アルミ押えバー","間仕切ネットBOXバー","落し","その他"]
    df_parts = {sn: R(sn) for sn in parts_sheets}
    df_gen = read_xlsx("原反価格表.xlsx")
    if df_gen.empty: df_gen = R("原反価格")
    return (df_clients, df_curtain, df_perf,
            df_ma, df_mb_tbl, df_mc, df_me_curt, df_me_motor,
            df_op, df_gap, df_parts, df_gen)

# ===== S・MAC 計算 =====
def extract_thickness(text: str) -> float | None:
    m = re.search(r"(\d+(?:\.\d+)?)\s*t", str(text or ""))
    if m:
        try: return float(m.group(1))
        except: return None
    return None

def smac_estimate(middle_name: str, open_method: str, W: int, H: int, cnt: int,
                  df_gen: pd.DataFrame, df_op: pd.DataFrame, picked_ops: list[dict]):
    """S・MAC販売額と、実使用のカーテン寸法（mm）を返す"""
    HEM_UNIT_THIN = 450
    HEM_UNIT_THICK = 550
    SEAM_UNIT_PER_M = 300

    res = {"ok": False, "msg": "", "sell_one": 0, "sell_total": 0,
           "note_ops": [], "breakdown": {}, "curtain_w": None, "curtain_h": None}
    if not middle_name or W<=0 or H<=0 or cnt<=0 or df_gen.empty:
        res["msg"] = "S・MAC：中分類/寸法/数量/原反価格を確認してください。"
        return res

    name_col = pick_col(df_gen, ["製品名","品名","名称"]) or df_gen.columns[0]
    w_col    = pick_col(df_gen, ["原反幅(mm)","原反幅","幅","巾"]) or df_gen.columns[1]
    u_col    = pick_col(df_gen, ["単価","上代","価格","金額"]) or df_gen.columns[2]
    t_col    = pick_col(df_gen, ["厚み","厚さ","t"])

    hit = df_gen[df_gen[name_col]==middle_name]
    if hit.empty:
        hit = df_gen[df_gen[name_col].astype(str).str.contains(re.escape(middle_name), na=False)]
    if hit.empty:
        res["msg"] = f"S・MAC：原反価格に『{middle_name}』が見つかりません。"
        return res

    gen_width = parse_float(hit.iloc[0][w_col])
    gen_price = parse_float(hit.iloc[0][u_col])
    thick = extract_thickness(hit.iloc[0][t_col]) if t_col else extract_thickness(hit.iloc[0][name_col])
    if not gen_width or not gen_price:
        res["msg"] = "S・MAC：原反幅または単価が不正です。"
        return res

    if open_method == "片引き":
        cur_w = W * 1.05; panels = 1
    else:
        cur_w = (W/2) * 1.05; panels = 2
    cur_h = H + 50
    res["curtain_w"] = int(round(cur_w))
    res["curtain_h"] = int(round(cur_h))

    length_per_panel_m = (cur_h * 1.2) / 1000.0
    joints = math.ceil(cur_w / gen_width)
    raw_len_m = length_per_panel_m * joints * panels
    raw_one = gen_price * raw_len_m

    cutting_one = (2000 if joints <= 3 else 3000) * panels
    seams_total = max(0, joints - 1) * panels
    seam_one = math.ceil(cur_h/1000.0) * SEAM_UNIT_PER_M * seams_total
    hem_unit = HEM_UNIT_THIN if (thick is not None and thick <= 0.3) else HEM_UNIT_THICK
    hem_perimeter_m = (cur_w + cur_w + cur_h + cur_h) / 1000.0
    fourfold_one = math.ceil(hem_perimeter_m) * hem_unit * panels

    note_ops, op_total = [], 0
    for op in (picked_ops or []):
        name = normalize_string(op.get("OP名称","")); unit = int(parse_float(op.get("金額")) or 0)
        dire = normalize_string(op.get("方向","")).upper()
        if not name or unit<=0: continue
        base_mm = cur_w if dire in ["W","横","X"] else cur_h
        units_1000 = math.ceil(base_mm/1000.0)
        sub = units_1000 * unit * panels * cnt
        op_total += sub; note_ops.append(name)
    op_one = op_total / cnt if cnt else 0

    genka_one   = raw_one + cutting_one + seam_one + fourfold_one + op_one
    genka_total = raw_one*cnt + cutting_one*cnt + seam_one*cnt + fourfold_one*cnt + op_total
    sell_one    = ceil100(genka_one / 0.6)
    sell_total  = ceil100(genka_total / 0.6)

    res.update({
        "ok": True,
        "sell_one": int(sell_one),
        "sell_total": int(sell_total),
        "note_ops": note_ops,
        "breakdown": {
            "原反材料(1式)":        int(round(raw_one)),
            "裁断賃(1式)":          int(round(cutting_one)),
            "幅繋ぎ(1式)":          int(round(seam_one)),
            "四方折り返し(1式)":    int(round(fourfold_one)),
            "OP加算(1式)":          int(round(op_one)),
            "原価(1式)":            int(round(genka_one)),
        }
    })
    return res

# ===== エア・セーブ／部材ヘルパ =====
def pick_price_col(df: pd.DataFrame):
    for c in ["㎡単価","平米単価","m2単価","単価","上代","価格","金額","固定価格"]:
        if c in df.columns: return c
    return None

def area_price(df_area: pd.DataFrame, item_name: str, perf: str, W: int, H: int, CNT: int):
    res = {"ok": False, "msg": "", "unit":0, "price_one":0, "total":0}
    if df_area.empty or not item_name or W<=0 or H<=0 or CNT<=0:
        res["msg"] = "単価表・品名・寸法・数量を確認してください。"; return res
    name_col = pick_col(df_area, ["品名","製品名","名称","品番","型式"]) or df_area.columns[0]
    perf_col = pick_col(df_area, ["性能","特性"])
    unit_col = pick_price_col(df_area)
    if unit_col is None: res["msg"] = "単価列が見つかりません。"; return res

    rows = df_area[(df_area[name_col]==item_name) & ((df_area[perf_col]==perf) if perf_col and perf else True)]
    if rows.empty: res["msg"] = "対象行が見つかりません。"; return res

    unit = parse_float(rows.iloc[0][unit_col])
    if not unit: res["msg"] = "単価が数値化できません。"; return res

    area = (W/1000)*(H/1000)
    price_one = ceil100(area*unit); total = int(price_one*CNT)
    res.update({"ok":True, "unit": int(round(unit)), "price_one": int(price_one), "total": int(total)})
    return res

def mc_slide_rail_price(W: int, CNT: int) -> int:
    rail_len_mm = int(math.ceil((W*2)/2000.0)*2000)
    return int((rail_len_mm/1000.0)*7400)*CNT

def pick_price_col_fixed(df_fixed: pd.DataFrame):
    return pick_col(df_fixed, ["固定価格","単価","価格","金額"]) or pick_price_col(df_fixed)

def fixed_price(df_fixed: pd.DataFrame, item_name: str) -> int:
    if df_fixed.empty or not item_name: return 0
    name_col = pick_col(df_fixed, ["品名","製品名","名称"]) or df_fixed.columns[0]
    price_col = pick_price_col_fixed(df_fixed)
    if price_col is None: return 0
    row = df_fixed[df_fixed[name_col]==item_name]
    if row.empty: return 0
    v = parse_float(row.iloc[0][price_col])
    return int(ceil100(v)) if v else 0

# ---- ME型OPの分類（カテゴリ列が無い場合のキーワード分類にも対応） ----
def split_me_op(df_me_op: pd.DataFrame):
    if df_me_op.empty: 
        return [], []
    name_col = pick_col(df_me_op, ["品名","製品名","名称"]) or df_me_op.columns[0]
    cat_col  = pick_col(df_me_op, ["カテゴリ","区分","種別","category"])
    names = df_me_op[name_col].dropna().astype(str).unique().tolist()

    motor_kw = ("駆動","ﾄﾞﾗｲﾌﾞ","モータ","ﾓｰﾀ","電動","駆動部","アクチュエータ","アクチュエーター")
    sensor_kw = ("センサー","sensor","ｾﾝｻ","インテリ","intelli","人感","リモコン","remote","コントローラ","制御","制御盤","スイッチ")

    motors, accs = [], []
    for nm in names:
        row = df_me_op[df_me_op[name_col]==nm].iloc[0]
        if cat_col:
            cat = str(row[cat_col]).lower()
            if any(k in cat for k in ["駆動","motor","モータ","電動"]):
                motors.append(nm); continue
            if any(k in cat for k in ["sensor","センサ","ｾﾝｻ","インテリ","制御","スイッチ","remote"]):
                accs.append(nm); continue
        low = nm.lower()
        if any(k.lower() in low for k in motor_kw):
            motors.append(nm)
        elif any(k.lower() in low for k in sensor_kw):
            accs.append(nm)
        else:
            accs.append(nm)
    if not motors and names:
        accs = names
    return sorted(set(motors)), sorted(set(accs))

PART_LENGTH_RULE = {
    "カーテンレール": "width",
    "取手付間仕切ポール": "height",
    "中間ポール": "height",
    "アルミ押えバー": "height",
    "間仕切ネットBOXバー": "height",
}
REQ_COLS_MAP = {
    "品名": ["品名","製品名","名称","品番"],
    "m単価": ["m単価","/m","1m単価","１m単価","m当たり","mあたり"],
    "セット長(mm)": ["セット長(mm)","セット長","定尺","規格長","サイズ","寸法","長さ","長さ(mm)"],
    "セット価格": ["セット価格","セット金額"],
    "固定価格": ["固定価格","単価","価格","金額","上代"],
}
def get_length_columns(df: pd.DataFrame) -> list[int]:
    lens = []
    for c in df.columns:
        try:
            v = int(str(c).replace(",", ""))
            if v>=1000: lens.append(v)
        except: pass
    return sorted(set(lens))

def pick_len_col(length_cols: list[int], need_mm: int) -> int | None:
    if not length_cols: return None
    for L in length_cols:
        if L>=need_mm: return L
    return length_cols[-1]

def map_required_cols(df: pd.DataFrame) -> dict:
    mp = {}
    for formal, cands in REQ_COLS_MAP.items():
        col = pick_col(df, cands)
        if col: mp[formal] = col
    return mp

def price_part_row(part_name: str, row: pd.Series, W: int, H: int, df_sheet: pd.DataFrame) -> tuple[int, str, float]:
    length_cols = get_length_columns(df_sheet)
    if length_cols:
        basis = 0
        if PART_LENGTH_RULE.get(part_name)=="width" and W>0:  basis=int(W)
        elif PART_LENGTH_RULE.get(part_name)=="height" and H>0: basis=int(H)
        if basis>0:
            col_mm = pick_len_col(length_cols, basis)
            if col_mm is not None:
                price_cell = row.get(col_mm) if col_mm in row else row.get(str(col_mm))
                price_raw = parse_float(price_cell)
                if price_raw and price_raw>0:
                    unit = ceil100(price_raw*0.5)
                    label = f"{int(col_mm/1000)}m相当"
                    return int(unit), label, float(col_mm)/1000.0
    mp = map_required_cols(pd.DataFrame([row]))
    if "m単価" in mp:
        per_m = parse_float(row[mp["m単価"]])
        if per_m and per_m>0:
            if PART_LENGTH_RULE.get(part_name)=="width" and W>0:
                mcount = math.ceil(W/1000.0)
            elif PART_LENGTH_RULE.get(part_name)=="height" and H>0:
                mcount = math.ceil((H/1000.0)*2)/2
            else: mcount = 0
            unit = int(ceil100(per_m*0.5) * (mcount if mcount else 0))
            label = f"{mcount:g}m分" if mcount else ""
            return unit, label, float(mcount or 0)
    if "セット長(mm)" in mp and "セット価格" in mp:
        L = parse_length_mm(row[mp["セット長(mm)"]]); P = parse_float(row[mp["セット価格"]])
        if L and L>=1000 and P and P>0:
            unit = ceil100(P*0.5); label=f"{int(L/1000)}mセット"
            return int(unit), label, float(L)/1000.0
    if "固定価格" in mp:
        P = parse_float(row[mp["固定価格"]])
        if P and P>0: return int(ceil100(P*0.5)), "", 0.0
    return 0, "", 0.0

# ===== 見積番号（3709-xxxxx 固定, 再生成ボタンなし） =====
def generate_estimate_no(seen_serials: set[str]) -> str:
    prefix = "3709"
    while True:
        sfx = f"{secrets.randbelow(100000):05d}"
        candidate = f"{prefix}-{sfx}"
        if candidate not in seen_serials:
            seen_serials.add(candidate); return candidate

# ===== 画面セットアップ =====
st.set_page_config(layout="wide", page_title="お見積書作成システム")

# CSS
st.markdown("""
<style>
section.main > div.block-container { padding-top: 8px; padding-bottom: 10px; padding-left: 10px; padding-right: 10px; }
div[data-testid="stHorizontalBlock"] { gap: 8px !important; }
div[data-testid="column"] { padding-left: 6px; padding-right: 6px; }
.sec-h { font-size:16pt; font-weight:400; margin: 6px 0 4px; }
.row-compact { display: grid; grid-template-columns: 1.4fr 0.45fr 0.7fr 0.8fr 0.28fr; column-gap: 8px; align-items: end; }
.sticky-wrap { position: sticky; top: 0; z-index: 999; background: var(--background-color); padding: 6px 6px 8px; border-bottom: 1px solid rgba(128,128,128,.35); }
.hr-thin { border-top: 1px solid rgba(128,128,128,.35); margin: 6px 0 10px; }
</style>
""", unsafe_allow_html=True)
def sec_title(text: str): st.markdown(f'<div class="sec-h">{text}</div>', unsafe_allow_html=True)

# マスタ
(df_clients, df_curtain, df_perf,
 df_ma, df_mb_tbl, df_mc, df_me_curt, df_me_motor,
 df_op, df_gap, df_parts, df_gen) = load_master()

# セッション
if "seen_serials" not in st.session_state: st.session_state.seen_serials = set()
if "estimate_base" not in st.session_state: st.session_state.estimate_base = generate_estimate_no(st.session_state.seen_serials)
if "openings" not in st.session_state:    st.session_state.openings = [{"id":1}]
if "file_title_manual" not in st.session_state: st.session_state.file_title_manual = False
if "file_title" not in st.session_state:        st.session_state.file_title = datetime.today().strftime("%m%d")

# ===== ヘッダ =====
st.markdown('<div class="sticky-wrap">', unsafe_allow_html=True)
sec_title("お見積書作成システム")

d1,d2,d3,d4 = st.columns([1.1,1.1,1.1,1.1])
clients = df_clients["社名"].dropna().unique().tolist() if "社名" in df_clients.columns else []
with d1:
    sel_client = st.selectbox("得意先名", [""]+clients, key="client")
with d2:
    branches = df_clients[df_clients["社名"]==sel_client]["支店名"].dropna().unique().tolist() if sel_client and "支店名" in df_clients.columns else []
    sel_branch = st.selectbox("支店名", [""]+branches, key="branch")
with d3:
    offices = df_clients[(df_clients["社名"]==sel_client)&(df_clients["支店名"]==sel_branch)]["営業所名"].dropna().unique().tolist() if sel_client and sel_branch and "営業所名" in df_clients.columns else []
    sel_office = st.selectbox("営業所名", [""]+offices, key="office")
with d4:
    person = st.text_input("担当者（得意先・全角ひらがな）", key="customer_pic")
    if person and not re.match(r"^[\u3041-\u3096ー\s]+$", person):
        st.error("担当者は全角ひらがなで入力してください。")

k1,k2,k3 = st.columns([1.2,1.0,0.8])
with k1:
    pj_name = st.text_input("物件名", key="pj")
    if not st.session_state.file_title_manual:
        st.session_state.file_title = f"{datetime.today().strftime('%m%d')}{pj_name}" if pj_name else datetime.today().strftime('%m%d')
with k2:
    our_staff_code = st.text_input("弊社担当者番号（半角英数2〜3）", key="our_staff_code")
    if our_staff_code and not re.match(r"^[A-Za-z0-9]{2,3}$", our_staff_code):
        st.error("弊社担当者番号は半角英数2〜3桁で入力してください。")
with k3:
    composed_no = f"{our_staff_code}-{st.session_state.estimate_base}" if our_staff_code else st.session_state.estimate_base
    st.text_input("見積番号", value=composed_no, key="estimate_no_disp", disabled=True)

m1 = st.columns([1.0])[0]
with m1:
    st.text_input("作成日", value=date.today().strftime("%Y/%m/%d"), key="created_disp", disabled=True)

st.markdown('<div class="hr-thin"></div>', unsafe_allow_html=True)
st.markdown("</div>", unsafe_allow_html=True)

# ===== サマリ用集計器 =====
overall_items = []   # 画面表示用の全明細（寸法行なども含む）
overall_total = 0
def overall_total_update(v): 
    global overall_total
    overall_total += int(v or 0)

# ===== 間口UI＋見積ロジック =====
PHRASES = [
    "金具：スチール", "金具：ステンレス",
    "戸先側：取手付間仕切りポール", "戸尻側：フラットバー固定", "戸尻側：取手付間仕切りポール", "戸尻側：吊下げ固定",
    "※ シートは収縮を考慮し長めでの出荷となります。現場で裾カット調整してください。",
    "※ カーテン下端は長めでの出荷となります。現場で裾カット調整してください。",
    "※ 下地別途。取付用のビス等は別途。", 
    "※ カーテンと間仕切りポール・中間ポールは組込済みです。取手・落し・マグネットは現場で取り付けてください。",
]

def render_opening(idx: int):
    pref = f"o{idx}_"

    a1,a2,a3,a4 = st.columns([0.7,1.0,1.0,0.7])
    with a1: mark = st.text_input("符号", key=pref+"mark")
    with a2: W = st.number_input("間口W (mm)", min_value=0, value=0, step=50, key=pref+"w")
    with a3: H = st.number_input("間口H (mm)", min_value=0, value=0, step=50, key=pref+"h")
    with a4: CNT = st.number_input("数量", min_value=1, value=1, step=1, key=pref+"cnt")

    sec_title("カーテン入力")
    b1,b2,b3 = st.columns([1.0,1.0,1.0])
    with b1:
        large_list = []
        if "大分類" in df_curtain.columns:
            large_list = [x for x in df_curtain["大分類"].dropna().unique().tolist() if x]
        if "エア・セーブ" not in large_list:
            large_list.append("エア・セーブ")
        large = st.radio("カーテン大分類", [""]+large_list, key=pref+"large", horizontal=True)

    air_type = None
    middle = small = perf = ""
    rib_note = ""

    with b2:
        if large and large != "エア・セーブ":
            mids = df_curtain[df_curtain["大分類"]==large]["中分類"].dropna().unique().tolist() if "中分類" in df_curtain.columns else []
            middle = st.selectbox("カーテン中分類", [""]+mids, key=pref+"mid")

    with b3:
        if large == "エア・セーブ":
            air_label = st.radio("型式（MA・MB・MC・ME）", ["","MA型折りたたみ式","MB型固定式","MC型スライド式","ME型電動式"], key=pref+"airtype", horizontal=True)
            air_type = air_label[:2] if air_label else None
        else:
            if middle:
                small_opts = df_curtain[(df_curtain["大分類"]==large)&(df_curtain["中分類"]==middle)]["小分類"].dropna().unique().tolist() if "小分類" in df_curtain.columns else []
                small = st.selectbox("カーテン小分類", [""]+small_opts, key=pref+"small")

    c1,c2,c3,c4 = st.columns([0.8,0.8,1.2,1.0])
    with c1:
        if large in ["S・MACカーテン"] or (large=="エア・セーブ" and air_type in ["MC","ME"]):
            open_mtd = st.radio("片引き/引分け", ["","片引き","引分け"], key=pref+"open", horizontal=True)
        else:
            open_mtd = ""

    with c2:
        if large=="エア・セーブ" and air_type in ["MB","MC","ME"]:
            rib_note = st.radio("リブ付き/リブ無し", ["","リブ付き","リブ無し"], key=pref+"rib", horizontal=True)

    with c3:
        air_item = ""
        if large=="エア・セーブ" and air_type:
            if air_type=="MA":
                name_col = pick_col(df_ma, ["品名","製品名","名称","品番","型式"]) or (df_ma.columns[0] if not df_ma.empty else None)
                items = df_ma[name_col].dropna().unique().tolist() if name_col else []
                air_item = st.selectbox("エア・セーブ品名", [""]+items, key=pref+"ma_item")
            elif air_type=="MB":
                name_col = pick_col(df_mb_tbl, ["品名","製品名","名称","品番","型式"]) or (df_mb_tbl.columns[0] if not df_mb_tbl.empty else None)
                items = df_mb_tbl[name_col].dropna().unique().tolist() if name_col else []
                air_item = st.selectbox("エア・セーブ品名", [""]+items, key=pref+"mb_item")
            elif air_type=="MC":
                name_col = pick_col(df_mc, ["品名","製品名","名称","品番","型式"]) or (df_mc.columns[0] if not df_mc.empty else None)
                items = df_mc[name_col].dropna().unique().tolist() if name_col else []
                air_item = st.selectbox("エア・セーブ品名", [""]+items, key=pref+"mc_item")
            elif air_type=="ME":
                curt_df = df_me_curt.copy() if not df_me_curt.empty else df_mc.copy()
                curt_name_col = pick_col(curt_df, ["品名","製品名","名称","品番","型式"]) or (curt_df.columns[0] if not df_curtain.empty else None)
                items = curt_df[curt_name_col].dropna().unique().tolist() if curt_name_col else []
                air_item = st.selectbox("エア・セーブ品名（カーテン）", [""]+items, key=pref+"me_curt")

    with c4:
        if large=="エア・セーブ":
            perf_all = df_perf["性能"].dropna().unique().tolist() if "性能" in df_perf.columns else []
            perf = st.selectbox("カーテン性能", [""]+perf_all, key=pref+"perf")
        elif large and large!="S・MACカーテン":
            perf_opts = df_perf[df_perf["中分類"]==middle]["性能"].dropna().unique().tolist() if "中分類" in df_perf.columns else []
            perf = st.selectbox("カーテン性能", [""]+perf_opts, key=pref+"perf2")

    picked_ops = []
    if large=="S・MACカーテン" and not df_op.empty and all(c in df_op.columns for c in ["OP名称","金額","方向"]):
        st.caption("S・MAC OP（任意／金額はカーテンに内包）")
        names = df_op["OP名称"].dropna().unique().tolist()
        cols = st.columns(3)
        for i, nm in enumerate(names):
            with cols[i%3]:
                if st.checkbox(nm, key=pref+f"smac_op_{i}"):
                    picked_ops.append(df_op[df_op["OP名称"]==nm].iloc[0].to_dict())

    # ===== 計算→ overall_items へ格納 =====
    if W>0 and H>0 and CNT>0:
        if large=="S・MACカーテン":
            sm = smac_estimate(middle or "", st.session_state.get(pref+"open") or "片引き",
                               W, H, CNT, df_gen, df_op, picked_ops)
            if sm["ok"]:
                # 行1: 品名（数量/単価は空）
                title = "S・MACカーテン"
                if middle: title += f" {middle}"
                if st.session_state.get(pref+'open'): title += f" {st.session_state.get(pref+'open')}"
                overall_items.append({
                    "opening": idx,
                    "品名": title,
                    "数量": "", "単位": "", "単価": "", "小計": 0,
                    "種別": "S・MAC",
                    "備考": (f"符号:{mark}" if mark else "")
                })
                # 行2: 間口寸法（空欄）
                overall_items.append({
                    "opening": idx,
                    "品名": f"間口寸法：W{W}×H{H}mm",
                    "数量": "", "単位": "", "単価": "", "小計": 0,
                    "種別": "寸法",
                    "備考": ""
                })
                # 行3: カーテン寸法（ここに数量/単価/小計）
                overall_items.append({
                    "opening": idx,
                    "品名": f"カーテン寸法：W{sm['curtain_w']}×H{sm['curtain_h']}mm",
                    "数量": CNT, "単位": "式",
                    "単価": sm["sell_one"], "小計": sm["sell_total"],
                    "種別": "寸法",
                    "備考": ("OP：" + "・".join(sm["note_ops"])) if sm["note_ops"] else ""
                })
                overall_total_update(sm["sell_total"])

                # UIのみの原価明細
                with st.expander(f"間口 {idx}：S・MAC 原価明細（UI表示のみ・Excel非出力）", expanded=False):
                    bd = sm["breakdown"]
                    df_bd = pd.DataFrame([
                        {"項目":"原反材料(1式)","金額":bd.get("原反材料(1式)",0)},
                        {"項目":"裁断賃(1式)","金額":bd.get("裁断賃(1式)",0)},
                        {"項目":"幅繋ぎ(1式)","金額":bd.get("幅繋ぎ(1式)",0)},
                        {"項目":"四方折り返し(1式)","金額":bd.get("四方折り返し(1式)",0)},
                        {"項目":"OP加算(1式)","金額":bd.get("OP加算(1式)",0)},
                        {"項目":"原価(1式)","金額":bd.get("原価(1式)",0)},
                    ])
                    df_bd["金額"] = df_bd["金額"].map(lambda x: f"¥{x:,}")
                    st.table(df_bd)
            else:
                st.warning(sm.get("msg") or "S・MACの計算に失敗しました。")

        elif large=="エア・セーブ" and air_type:
            # 共通ヘルパ：品名行＋間口寸法行（価格は寸法行へ）
            def push_two_lines(title, price_one, total, note_extra=""):
                # 行1: 品名（空）
                overall_items.append({
                    "opening": idx,
                    "品名": title,
                    "数量": "", "単位":"", "単価":"", "小計": 0,
                    "種別": f"エア・セーブ{air_type}",
                    "備考": (f"符号:{mark}" if mark else "")
                })
                # 行2: 間口寸法（ここに価格）
                overall_items.append({
                    "opening": idx,
                    "品名": f"間口寸法：W{W}×H{H}mm",
                    "数量": CNT, "単位":"式", "単価": price_one, "小計": total,
                    "種別": "寸法", "備考": note_extra
                })
                overall_total_update(total)

            if air_type=="MA" and air_item and perf:
                r = area_price(df_ma, air_item, perf, W, H, CNT)
                if r["ok"]:
                    title = f"エア・セーブ MA型折りたたみ式 {air_item}"
                    push_two_lines(title, r["price_one"], r["total"])

            elif air_type=="MB" and air_item and perf:
                r = area_price(df_mb_tbl, air_item, perf, W, H, CNT)
                if r["ok"]:
                    title = f"エア・セーブ MB型固定式 {air_item}"
                    note = (f"{st.session_state.get(pref+'rib')}" if st.session_state.get(pref+'rib') else "")
                    push_two_lines(title, r["price_one"], r["total"], note)

            elif air_type=="MC" and air_item and perf:
                r = area_price(df_mc, air_item, perf, W, H, CNT)
                if r["ok"]:
                    # カーテン本体（2行構成）
                    title = f"エア・セーブ MC型スライド式 {air_item}"
                    note = " / ".join([x for x in [
                        st.session_state.get(pref+'rib'),
                        st.session_state.get(pref+'open')
                    ] if x])
                    push_two_lines(title, r["price_one"], r["price_one"]*CNT, note)
                    # レールは従来通り単独行（寸法行は付けない）
                    rail = mc_slide_rail_price(W, CNT)
                    overall_items.append({
                        "opening": idx,
                        "品名": "スライドレール", "数量": 1, "単位":"式", "単価": rail, "小計": rail,
                        "種別":"エア・セーブMC", "備考": "W×2を2000mm刻み"
                    })
                    overall_total_update(rail)

            elif air_type=="ME" and perf:
                total_me = 0
                # ① カーテン（2行構成）
                if air_item:
                    curt_df = df_me_curt.copy() if not df_me_curt.empty else df_mc.copy()
                    r = area_price(curt_df, air_item, perf, W, H, CNT)
                    if r["ok"]:
                        title = f"エア・セーブ ME型電動式 カーテン {air_item}"
                        overall_items.append({
                            "opening": idx,
                            "品名": title,
                            "数量": "", "単位":"", "単価":"", "小計": 0,
                            "種別":"エア・セーブME(カーテン)",
                            "備考": (f"符号:{mark}" if mark else "")
                        })
                        overall_items.append({
                            "opening": idx,
                            "品名": f"間口寸法：W{W}×H{H}mm",
                            "数量": CNT, "単位":"式", "単価": r["price_one"], "小計": r["total"],
                            "種別":"寸法", "備考": ""
                        })
                        total_me += r["total"]

                # ② ME型OP（駆動部：単一選択／センサー等：複数選択）
                motors, accs = split_me_op(df_me_motor)
                me_motor_sel = ""
                if motors:
                    me_motor_sel = st.selectbox("ME：電動駆動部（必須1つ）", [""]+motors, key=pref+"me_motor")
                acc_selected = []
                if accs:
                    st.caption("ME：インテリジェントセンサー／リモコン／制御ほか（複数選択可）")
                    acc_cols = st.columns(3)
                    for i, nm in enumerate(accs):
                        with acc_cols[i%3]:
                            if st.checkbox(nm, key=pref+f"me_acc_{i}"):
                                acc_selected.append(nm)

                if me_motor_sel:
                    mu = fixed_price(df_me_motor, me_motor_sel)
                    if mu>0:
                        overall_items.append({
                            "opening": idx,
                            "品名": f"エア・セーブ ME型電動式 駆動部 {me_motor_sel}",
                            "数量": CNT, "単位":"式", "単価": mu, "小計": mu*CNT,
                            "種別":"エア・セーブME(駆動部)", "備考": ""
                        })
                        total_me += mu*CNT

                if acc_selected:
                    for nm in acc_selected:
                        pv = fixed_price(df_me_motor, nm)
                        if pv>0:
                            overall_items.append({
                                "opening": idx,
                                "品名": f"エア・セーブ ME型OP {nm}",
                                "数量": CNT, "単位":"式", "単価": pv, "小計": pv*CNT,
                                "種別":"エア・セーブME(OP)", "備考": ""
                            })
                            total_me += pv*CNT

                if total_me>0:
                    overall_total_update(total_me)

    # ===== 部材入力（S・MAC または エア・セーブMAで表示） =====
    show_parts = (
        st.session_state.get(pref+"large")=="S・MACカーテン" or
        (st.session_state.get(pref+"large")=="エア・セーブ" and (st.session_state.get(pref+"airtype") or "").startswith("MA"))
    )
    if show_parts:
        st.markdown("##### 部材入力")
        for sheet_name, dfp in df_parts.items():
            rows_key = pref + f"{sheet_name}_rows"
            if rows_key not in st.session_state:
                st.session_state[rows_key] = [{"item":"", "qty":1}]

            # 「＋」を見出しの左側へ
            hcolL, hcolR = st.columns([0.14, 0.86])
            with hcolL:
                if st.button("＋", key=pref+f"{sheet_name}_add"):
                    st.session_state[rows_key].append({"item":"", "qty":1}); st.rerun()
            with hcolR:
                st.caption(f"【{sheet_name}】")

            name_col = pick_col(dfp, ["品名","製品名","名称","品番"]) or (dfp.columns[0] if not dfp.empty else None)
            names = dfp[name_col].dropna().unique().tolist() if name_col else []
            updated_rows = []
            for j, rowdata in enumerate(st.session_state[rows_key]):
                st.markdown('<div class="row-compact">', unsafe_allow_html=True)
                col1, col2, col3, col4, col5 = st.columns([1.4,0.45,0.7,0.8,0.28])
                with col1:
                    all_item_opts = names
                    current = rowdata["item"] if rowdata["item"] in ([""] + all_item_opts) else ""
                    item_opts = [""] + ([current] if current else all_item_opts)
                    try:
                        default_idx = item_opts.index(current) if current in item_opts else 0
                    except ValueError:
                        default_idx = 0
                    sel = st.radio(
                        f"品名 {j+1}",
                        item_opts,
                        index=default_idx,
                        key=pref+f"{sheet_name}_item_{j}_radio"
                    )
                with col2:
                    qty = st.number_input("数量", min_value=1, value=int(rowdata["qty"]), step=1, key=pref+f"{sheet_name}_qty_{j}")
                unit, label, used_m = (0,"",0.0)
                if sel:
                    src_row = dfp[dfp[name_col]==sel].iloc[0] if not dfp.empty else None
                    if src_row is not None:
                        unit, label, used_m = price_part_row(sheet_name, src_row, W, H, dfp)
                with col3: st.text_input("単価（自動）", value=str(unit), disabled=True, key=pref+f"{sheet_name}_unit_{j}")
                with col4: st.text_input("使用m数（自動）", value=(f"{used_m:g}" if used_m else ""), disabled=True, key=pref+f"{sheet_name}_m_{j}")
                with col5: delete = st.button("×", key=pref+f"{sheet_name}_del_{j}")
                st.markdown('</div>', unsafe_allow_html=True)

                if not delete:
                    updated_rows.append({"item": sel, "qty": qty})
                    if sel and unit>0:
                        subtotal = unit*qty
                        overall_items.append({
                            "opening": idx,
                            "品名": sel, "数量": qty, "単位":"式", "単価": unit, "小計": subtotal,
                            "種別":"部材", "備考": (f"符号:{mark}" if mark else "") + (f"／{label}" if label else "")
                        })
                        overall_total_update(subtotal)
            st.session_state[rows_key] = updated_rows

        # 定型文（間口末尾A列へ1行で出力）: 間口ごと
        st.markdown("##### 定型文")
        phr_sel = []
        colL, colR = st.columns(2)
        half = (len(PHRASES)+1)//2
        with colL:
            for i, p in enumerate(PHRASES[:half]):
                if st.checkbox(p, key=pref+f"phr_L_{i}"):
                    phr_sel.append(p)
        with colR:
            for i, p in enumerate(PHRASES[half:]):
                if st.checkbox(p, key=pref+f"phr_R_{i}"):
                    phr_sel.append(p)
        if phr_sel:
            for p in phr_sel:
                overall_items.append({
                    "opening": idx,
                    "品名": "（定型文）",
                    "数量": "", "単位":"", "単価":"", "小計":0, "種別":"メモ",
                    "備考": p
                })
    else:
        st.caption("※このカーテン構成では部材の追加は行いません。")

# ===== 間口描画 =====
for i, op in enumerate(st.session_state.openings, start=1):
    with st.expander(f"間口 {i}", expanded=True):
        render_opening(i)
    cols = st.columns([0.16, 0.84])
    if cols[0].button("この間口を削除", key=f"del_{i}") and len(st.session_state.openings)>1:
        st.session_state.openings.pop(i-1); st.rerun()
if st.button("＋ 間口を追加", key="add_opening"):
    st.session_state.openings.append({"id": len(st.session_state.openings)+1}); st.rerun()

# ===== サマリ表示（画面） =====
st.markdown("---")
sec_title("見積サマリ")
if overall_items:
    df_summary = pd.DataFrame(overall_items, columns=["opening","品名","数量","単位","単価","小計","種別","備考"])
    st.dataframe(df_summary, use_container_width=True, hide_index=True)
else:
    st.info("明細がありません。間口を追加し、カーテン／部材を入力してください。")

sec_title("見積金額")
st.metric("税抜合計", f"¥{overall_total:,}")

# ===== 運賃・梱包 =====
st.markdown("---")
sec_title("運賃・梱包")
ship_method = st.radio("配送条件", ["","路線便（時間指定不可）","現場搬入（時間指定可）"], horizontal=True, key="ship_method")
ship_fee = st.number_input("金額", min_value=0, step=100, key="ship_fee")

# ===== 開口別集計（見積書0用） =====
def split_by_opening(items: list[dict]):
    by_op = {}
    for it in items:
        op = it.get("opening", 0)
        by_op.setdefault(op, []).append(it)
    return [by_op[k] for k in sorted(by_op.keys())]

def opening_summary(opening_items: list[dict]):
    curtain_name = ""
    total = 0
    for it in opening_items:
        typ = it.get("種別","")
        # 最初に出てくる品名行（S・MAC/エア・セーブ）をA列名に使う
        if not curtain_name and (typ.startswith("S・MAC") or typ.startswith("エア・セーブ")):
            curtain_name = it.get("品名","")
        if typ != "メモ":
            total += int(it.get("小計") or 0)
    memos = [it.get("備考","") for it in opening_items if it.get("種別")=="メモ" and it.get("備考")]
    memo_text = " / ".join(memos) if memos else ""
    return curtain_name or "（カーテン未選択）", total, memo_text

# ===== Excel結合セル 安全書き込みヘルパ =====
def write_merged_safe(ws, coord: str, value):
    """結合セル範囲内であっても先頭セルに安全に書き込む"""
    r, c = coordinate_to_tuple(coord)
    target = coord
    for mr in ws.merged_cells.ranges:
        if (r, c) in mr.cells:
            top_left = mr.min_row, mr.min_col
            target = ws.cell(row=top_left[0], column=top_left[1]).coordinate
            break
    ws[target].value = value

# ===== バリデーション =====
def validate_before_export(items: list[dict], ship_method: str, ship_fee: int):
    errs = []
    if not items:
        errs.append("明細がありません。")
        return errs

    # 梱包必須：S・MAC / エア・セーブMA を含む場合
    has_required = any(it.get("種別") in ["S・MAC","エア・セーブMA"] for it in items)
    if has_required and (not ship_method or int(ship_fee)<=0):
        errs.append("運賃・梱包は必須です。配送条件と金額を入力してください。")

    # 見積書0行数（間口数 + 運賃行）≤ 24
    groups = split_by_opening(items)
    line_count = len(groups) + (1 if ship_method else 0)
    if line_count > 24:
        errs.append(f"見積書0の行数オーバー（21〜44行＝最大24行）。現在 {line_count} 行。間口を減らすかまとめてください。")

    # テンプレート存在＆必要シート
    need_sheets = ["見積書0","見積書1","見積書2","見積書3","見積書4","見積書5"]
    if not TEMPLATE_BOOK.exists():
        errs.append("テンプレートがありません：お見積書（明細）.xlsx")
    else:
        try:
            wb = load_workbook(TEMPLATE_BOOK)
            for sn in need_sheets:
                if sn not in wb.sheetnames:
                    errs.append(f"テンプレート（お見積書（明細）.xlsx）にシート『{sn}』がありません。")
        except Exception as e:
            errs.append(f"テンプレートを開けません: {e}")
    return errs

# ===== Excel出力 =====
def export_to_template(out_path: str, items: list[dict], header: dict, ship_method: str, ship_fee: int):
    wb = load_workbook(TEMPLATE_BOOK)
    ws0 = wb["見積書0"]

    # ヘッダー（結合セルでもOKな安全書き込み）
    write_merged_safe(ws0, "J1", header["estimate_no"])
    write_merged_safe(ws0, "J3", header["date"])
    write_merged_safe(ws0, "A6", header["customer_name"])
    write_merged_safe(ws0, "A7", (f"{header['branch_name']} {header['office_name']}".strip()))
    write_merged_safe(ws0, "A8", header["person_name"])
    write_merged_safe(ws0, "B17", header["project_name"])

    # 見積書0：21〜 に間口合計（A/F/G/H）→ 末尾に運賃梱包（空行なし）
    row = 21
    opening_groups = split_by_opening(items)
    for grp in opening_groups:
        name, total, _ = opening_summary(grp)
        ws0[f"A{row}"] = name
        ws0[f"F{row}"] = 1
        ws0[f"G{row}"] = f"=H{row}"
        ws0[f"H{row}"] = total
        row += 1

    if ship_method:
        ws0[f"A{row}"] = "運賃・梱包 " + ship_method
        ws0[f"F{row}"] = 1
        ws0[f"G{row}"] = f"=H{row}"
        ws0[f"H{row}"] = int(ship_fee)
        row += 1

    # 見積書1〜5：11行目は触らず、12〜44（33行/頁）
    def lines_for_opening(grp: list[dict]):
        lines = []
        # 通常行（品名・数量・単位・単価）…「メモ」以外はすべて出す（寸法行含む）
        for it in grp:
            if it.get("種別") != "メモ":
                lines.append((
                    it.get("品名",""),
                    it.get("数量",""),
                    it.get("単位",""),
                    it.get("単価",""),
                ))
        # メモ行（定型句は1項目=1行）
        for it in grp:
            if it.get("種別") == "メモ":
                txt = it.get("備考","")
                if txt:
                    lines.append((txt, "", "", ""))
        return lines

    blocks = [lines_for_opening(grp) for grp in opening_groups]
    # 間口間に空行を1つ挿入（最後以外）
    for bi in range(len(blocks)-1):
        blocks[bi].append(("", "", "", ""))

    PAGES = [wb[f"見積書{i}"] for i in range(1,6)]
    page_idx = 0
    row_in_page = 12

    def new_page():
        nonlocal page_idx, row_in_page
        page_idx += 1
        row_in_page = 12

    for op_lines in blocks:
        block_len = len(op_lines)
        if block_len == 0:
            continue
        # 1間口 ≤ 33行ならページ跨ぎ禁止（今のページに入らなければ改ページ）
        if block_len <= 33 and (row_in_page + block_len - 1) > 44:
            new_page()
        while op_lines:
            if page_idx >= len(PAGES):
                raise RuntimeError("明細ページが5ページを超えました。")
            ws = PAGES[page_idx]
            remain = 44 - row_in_page + 1
            take = min(len(op_lines), remain)
            chunk, op_lines = op_lines[:take], op_lines[take:]
            for a, f, g, h in chunk:
                ws[f"A{row_in_page}"] = a
                ws[f"F{row_in_page}"] = f
                ws[f"G{row_in_page}"] = g
                ws[f"H{row_in_page}"] = h
                row_in_page += 1
            if op_lines:
                new_page()

    used_pages = page_idx + (1 if row_in_page > 12 else 0)
    if used_pages > 5:
        raise RuntimeError(f"明細ページが5ページを超えました（使用{used_pages}ページ）。")

    os.makedirs(osp.dirname(out_path), exist_ok=True)
    wb.save(out_path)

# ===== 保存UI =====
st.markdown("---")
sec_title("保存")
save_dir = "./data"; os.makedirs(save_dir, exist_ok=True)

c1, c2 = st.columns([0.6, 0.4])
with c1:
    st.checkbox("保存ファイル名を手動で編集する", value=st.session_state.file_title_manual, key="file_title_manual")
    st.text_input("保存ファイル名", value=st.session_state.file_title, key="file_title", disabled=not st.session_state.file_title_manual)

with c2:
    st.markdown("**Excel保存（お見積書（明細）.xlsx へ直接転記）**")
    if st.button("Excel保存（テンプレ転記）", key="excel_save_btn"):
        errs = validate_before_export(overall_items, st.session_state.get("ship_method",""), int(st.session_state.get("ship_fee") or 0))
        if errs:
            for e in errs: st.error(e)
        else:
            try:
                header = {
                    "estimate_no":   (f"{st.session_state.get('our_staff_code')}-" if st.session_state.get('our_staff_code') else "") + st.session_state.get("estimate_base",""),
                    "date":          (st.session_state.get("created_disp") or datetime.today().strftime("%Y/%m/%d")).replace("/","-"),
                    "customer_name": st.session_state.get("client",""),
                    "branch_name":   st.session_state.get("branch",""),
                    "office_name":   st.session_state.get("office",""),
                    "person_name":   st.session_state.get("customer_pic",""),
                    "project_name":  st.session_state.get("pj",""),
                }
                out = osp.join(save_dir, f"{st.session_state.file_title}_見積書.xlsx")
                export_to_template(out, overall_items, header, st.session_state.get("ship_method",""), int(st.session_state.get("ship_fee") or 0))
                st.success(f"Excelを保存しました：{out}")
                with open(out, "rb") as f:
                    st.download_button("ダウンロード", f.read(),
                        file_name=os.path.basename(out),
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="excel_dl_btn")
            except Exception as e:
                st.error("Excel出力でエラーが発生しました。")
                st.exception(e)
